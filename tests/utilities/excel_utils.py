import openpyxl
from openpyxl.styles import PatternFill


# number of rows
def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


# number of columns
def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


# read data
def read_data(file, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row_num, column_num).value


# write data
def write_data(file, sheetname, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row_num, column_num).value = data
    workbook.save(file)
